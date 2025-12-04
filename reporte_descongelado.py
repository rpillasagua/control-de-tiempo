import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import numpy as np
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# --- Funciones de Interfaz de Usuario ---
def seleccionar_archivos_multiples():
    """Permite seleccionar múltiples archivos y detecta automáticamente el tipo"""
    root = tk.Tk()
    root.withdraw()
    
    archivos = filedialog.askopenfilenames(
        title="Selecciona los archivos de reporte (ENTERO y/o COLA)",
        filetypes=[("Archivos de Excel", "*.xlsx *.xlsm")]
    )
    
    if not archivos:
        return []
    
    archivos_clasificados = []
    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo).upper()
        
        if "ENTERO" in nombre_archivo:
            tipo = "ENTERO"
        elif "COLA" in nombre_archivo:
            tipo = "COLA"
        else:
            # Si no se puede detectar automáticamente, preguntar
            respuesta = messagebox.askyesno(
                "Clasificar Archivo",
                f"No se pudo detectar automáticamente el tipo del archivo:\n{nombre_archivo}\n\n¿Es un reporte de ENTERO?\n(No = COLA)"
            )
            tipo = "ENTERO" if respuesta else "COLA"
        
        archivos_clasificados.append((tipo, archivo))
        print(f"Detectado: {tipo} - {os.path.basename(archivo)}")
    
    return archivos_clasificados

def seleccionar_archivo_excel(titulo):
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(title=titulo, filetypes=[("Archivos de Excel", "*.xlsx *.xlsm")])

def seleccionar_ruta_guardado():
    root = tk.Tk()
    root.withdraw()
    return filedialog.asksaveasfilename(
        title="Guardar reporte 'Lotes con Novedades en Descongelado' como...", 
        defaultextension=".xlsx",
        initialfile="Lotes con Novedades en Descongelado.xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )

def obtener_fecha_por_reporte(tipo_reporte, fecha_maxima):
    """Solicita fecha de inicio específica para cada tipo de reporte"""
    print(f"\n--- Configuracion para {tipo_reporte} ---")
    print(f"Fecha mas reciente en {tipo_reporte}: {fecha_maxima.strftime('%d/%m/%y')}")
    
    while True:
        fecha_str = input(f"Ingrese fecha de inicio para {tipo_reporte} (dd/mm/aa): ")
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%d/%m/%y')
            if fecha_inicio > fecha_maxima:
                print(f"ADVERTENCIA: La fecha ingresada es posterior a la mas reciente ({fecha_maxima.strftime('%d/%m/%y')})")
                continuar = input("Desea continuar? (s/n): ").lower().startswith('s')
                if not continuar:
                    continue
            return fecha_inicio
        except ValueError: 
            print("Formato incorrecto. Use dd/mm/aa")

# --- Funciones de Validación ---
def safe_float(value, default=np.nan):
    if pd.isna(value): return np.nan
    try:
        return float(str(value).replace(',', '.'))
    except (ValueError, TypeError):
        return np.nan

def validar_peso_bruto(fila, info_producto, tipo_reporte):
    valor_reporte_g = safe_float(fila.iloc[9])
    if info_producto.empty or pd.isna(valor_reporte_g): return None, None
    
    try:
        embalaje_str = str(info_producto['EMBALAJE'].iloc[0])
        num_unidades = 1
        if "Und" in embalaje_str:
            partes = embalaje_str.split()
            if partes[0].isdigit():
                num_unidades = int(partes[0])
        
        unidad_medida = "KG"
        if embalaje_str.upper().endswith("LB"):
            unidad_medida = "LB"

        # LÓGICA CORREGIDA según la macro VBA:
        # Si numUnidades >= 3 usa columna R (PESO_BRUTO_PRODUCCION)
        # Si numUnidades < 3 usa columna T (PESO_BRUTO_MASTERS)
        if num_unidades >= 3:
            col_peso_base = 'PESO_BRUTO_PRODUCCION'  # Columna R
        else:
            col_peso_base = 'PESO_BRUTO_MASTERS'     # Columna T
        
        peso_base_valor = safe_float(info_producto[col_peso_base].iloc[0])
        peso_base_g = 0
        if unidad_medida == "LB":
            peso_base_g = peso_base_valor * 453.592
        else:
            peso_base_g = peso_base_valor * 1000

        glaseo_g = 0
        # LÓGICA CORREGIDA: Solo calcular glaseo si numUnidades >= 3
        if num_unidades >= 3:
            glaseo_valor = safe_float(info_producto['GLASEO_VALOR'].iloc[0])
            glaseo_unidad = str(info_producto['GLASEO_UNIDAD'].iloc[0])
            if glaseo_valor > 0:
                if glaseo_unidad == '%':
                    peso_neto_kg = safe_float(info_producto['PESO_NETO'].iloc[0])
                    glaseo_g = (peso_neto_kg * 1000) * (glaseo_valor / 100)
                else:
                    glaseo_g = glaseo_valor

        sobrepeso_porc = safe_float(str(info_producto['SOBREPESO_PORCENTAJE'].iloc[0]).replace('%', '')) / 100
        limite_inferior_g = peso_base_g + glaseo_g
        limite_superior_g = limite_inferior_g + (peso_base_g * sobrepeso_porc)

        if valor_reporte_g > limite_superior_g:
            return f"PESO BRUTO alto ({valor_reporte_g / 1000:.2f} Kg)", f"P. BRUTO FT: {limite_inferior_g/1000:.2f}-{limite_superior_g/1000:.2f} KG"
        if valor_reporte_g < limite_inferior_g:
            return f"PESO BRUTO bajo ({valor_reporte_g / 1000:.2f} Kg)", f"P. BRUTO FT: {limite_inferior_g/1000:.2f}-{limite_superior_g/1000:.2f} KG"

    except Exception: return None, None
    return None, None

def validar_peso_neto(fila, info_producto, tipo_reporte):
    valor_reporte_g = safe_float(fila.iloc[10])
    if info_producto.empty or pd.isna(valor_reporte_g): return None, None
    try:
        embalaje_str = str(info_producto['EMBALAJE'].iloc[0])
        unidad_medida = "KG"
        if embalaje_str.upper().endswith("LB"):
            unidad_medida = "LB"

        neto_base_valor = safe_float(info_producto['PESO_NETO'].iloc[0])
        neto_base_g = 0
        if unidad_medida == "LB":
            neto_base_g = neto_base_valor * 453.592
        else:
            neto_base_g = neto_base_valor * 1000

        sobrepeso_porc = safe_float(str(info_producto['SOBREPESO_PORCENTAJE'].iloc[0]).replace('%', '')) / 100
        limite_superior_g = neto_base_g * (1 + sobrepeso_porc)
        
        if valor_reporte_g > limite_superior_g:
            return f"PESO NETO alto ({valor_reporte_g / 1000:.2f} Kg)", f"P. NETO FT: {neto_base_g/1000:.2f}-{limite_superior_g/1000:.2f} KG"
        if valor_reporte_g < neto_base_g:
            return f"PESO NETO bajo ({valor_reporte_g / 1000:.2f} Kg)", f"P. NETO FT: {neto_base_g/1000:.2f}-{limite_superior_g/1000:.2f} KG"
    except Exception: return None, None
    return None, None

def validar_talla(fila, codigo_str, df_tallas, tipo_reporte):
    try:
        talla_ingresada = str(fila.iloc[8]).replace('/', '-')
        talla_en_ft = df_tallas[(df_tallas['CODIGO'] == codigo_str) & (df_tallas['TALLA_MARCADA'] == talla_ingresada)]
        if talla_en_ft.empty:
            return "TALLA NO EXISTE EN FT", f"TALLA: {talla_ingresada} no válida"
    except Exception: return None, None
    return None, None

def validar_uniformidad(fila, codigo_str, df_tallas, tipo_reporte):
    grandes = safe_float(fila.iloc[13])
    pequenos = safe_float(fila.iloc[14])
    if pd.isna(grandes) or pd.isna(pequenos) or pequenos == 0: return None, None
    try:
        ratio_calculado = round(grandes / pequenos, 2)
        talla_ingresada = str(fila.iloc[8]).replace('/', '-')
        talla_info = df_tallas[(df_tallas['CODIGO'] == codigo_str) & (df_tallas['TALLA_MARCADA'] == talla_ingresada)]
        if talla_info.empty: return None, None
        limite = safe_float(talla_info['UNIFORMIDAD'].iloc[0])
        if ratio_calculado > limite:
            return f"UNIFORMIDAD alta ({ratio_calculado})", f"UNIFORMIDAD MÁX: {limite}"
    except Exception: return None, None
    return None, None

def validar_conteo_por_kg(fila, codigo_str, df_tallas, tipo_reporte):
    valor_reporte = safe_float(fila.iloc[12])
    if pd.isna(valor_reporte): return None, None
    try:
        talla_ingresada = str(fila.iloc[8]).replace('/', '-')
        talla_info = df_tallas[(df_tallas['CODIGO'] == codigo_str) & (df_tallas['TALLA_MARCADA'] == talla_ingresada)]
        if talla_info.empty: return None, None
        rango_str = talla_info['CONTEO_FINAL'].iloc[0]
        limites = [safe_float(n) for n in re.findall(r'(\d+\.?\d*)', str(rango_str))]
        if len(limites) != 2: return None, None
        limite_inferior, limite_superior = min(limites), max(limites)
        if valor_reporte > limite_superior:
            return f"CONTEO alto ({int(valor_reporte)})", f"CONTEO FT: {rango_str}"
        if valor_reporte < limite_inferior:
            return f"CONTEO bajo ({int(valor_reporte)})", f"CONTEO FT: {rango_str}"
    except Exception: return None, None
    return None, None

def validar_defectos(fila, codigo_str, df_defectos, tipo_reporte):
    """Validar defectos - LÓGICA ORIGINAL RESTAURADA"""
    motivos, especificaciones = [], []
    conteo_caja = safe_float(fila.iloc[11])  # Columna L
    if pd.isna(conteo_caja) or conteo_caja == 0: return [], []
    
    # Mapeo EXACTO del código original
    if tipo_reporte == "ENTERO":
        mapa_defectos = [("CABEZA_ROJA", 16), ("CABEZA_NARANJA", 18), ("CABEZA_FLOJA", 20), ("CABEZA_DESCOLGADA", 22), ("BRANQUIAS_OSCURAS_LEVES", 24), ("BRANQUIAS_OSCURAS_FUERTES", 26), ("BRANQUIAS_AMARILLAS_LEVES", 28), ("BRANQUIAS_AMARILLAS_FUERTES", 30), ("HEMOLINFAS_LEVES", 32), ("HEMOLINFAS_FUERTES", 34), ("HEPATO_REGADO", 36), ("HEPATO_REVENTADO", 38), ("FLACIDO", 40), ("MUDADO", 42), ("MANCHAS_NEGRAS_LEVES", 44), ("MANCHAS_NEGRAS_FUERTES", 46), ("DEFORMES_LEVES", 48), ("DEFORMES_FUERTES", 50), ("QUEBRADO", 52), ("MELANOSIS", 54), ("MALTRATADO", 56), ("HONGO_BUCAL_LEVE", 58), ("HONGO_BUCAL_FUERTE", 60)]
    else:  # COLA  
        mapa_defectos = [("FLACIDO", 16), ("MUDADO", 18), ("MANCHAS_NEGRAS_LEVES", 20), ("MANCHAS_NEGRAS_FUERTES", 22), ("DEFORMES_LEVES", 24), ("DEFORMES_FUERTES", 26), ("DESHIDRATADO", 28), ("QUEBRADO", 30), ("MELANOSIS", 32), ("MALTRATADO", 34), ("MAL_DESCABEZADO", 36), ("SEMIROSADO", 38), ("ROSADO", 40), ("PATAS", 42), ("RESIDUOS_DE_HEPATOPANCREAS", 44)]
    
    suma_porcentajes = 0
    for _, col_idx in mapa_defectos:
        num_defectos = safe_float(fila.iloc[col_idx], 0)
        if num_defectos > 0:
            suma_porcentajes += (num_defectos / conteo_caja) * 100
    for nombre_defecto, col_idx in mapa_defectos:
        num_defectos = safe_float(fila.iloc[col_idx], 0)
        if num_defectos > 0:
            porcentaje_calculado = (num_defectos / conteo_caja) * 100
            limite_info = df_defectos[(df_defectos['CODIGO'] == codigo_str) & (df_defectos['DEFECTO'] == nombre_defecto)]
            if not limite_info.empty:
                limite_str = str(limite_info['VALOR'].iloc[0]).upper().strip()
                fuera_de_parametro = False
                if limite_str == "SI": continue
                if limite_str == "NO": fuera_de_parametro = True
                elif '%' in limite_str:
                    limite_num = safe_float(limite_str.replace('%', ''))
                    if not pd.isna(limite_num) and porcentaje_calculado > limite_num: fuera_de_parametro = True
                if fuera_de_parametro:
                    motivos.append(f"{nombre_defecto.replace('_', ' ')} alto ({porcentaje_calculado:.2f}%)")
                    especificaciones.append(f"{nombre_defecto.replace('_', ' ')} MÁX: {limite_str}")
    limite_total_info = df_defectos[(df_defectos['CODIGO'] == codigo_str) & (df_defectos['DEFECTO'] == 'DEFECTOS_TOTALES')]
    if not limite_total_info.empty:
        limite_total_str = str(limite_total_info['VALOR'].iloc[0])
        if '%' in limite_total_str:
            limite_total_num = safe_float(limite_total_str.replace('%', ''))
            if not pd.isna(limite_total_num) and suma_porcentajes > limite_total_num:
                motivos.append(f"TOTAL DEFECTOS alto ({suma_porcentajes:.2f}%)")
                especificaciones.append(f"TOTAL DEFECTOS MÁX: {limite_total_str}")
    return motivos, especificaciones

def format_excel_output_profesional(writer, df):
    """Formato profesional con encabezado azul y ajuste automático de texto"""
    worksheet = writer.sheets['Lotes con Novedades en Descongelado']
    
    # Colores y estilos
    azul_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    borde_delgado = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatear encabezados
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    for col_idx, cell in enumerate(worksheet[1], 1):
        cell.font = header_font
        cell.fill = azul_header
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = borde_delgado
    
    # Formatear datos
    data_font = Font(name='Calibri', size=10)
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = borde_delgado
            
            # Alternar colores de fila
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # AJUSTE AUTOMÁTICO DE COLUMNAS - MEJORADO
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        column_name = str(worksheet.cell(row=1, column=col_idx).value) if worksheet.cell(row=1, column=col_idx).value else ""
        
        # Calcular ancho basado en contenido
        max_length = 0
        
        # Revisar el contenido de todas las celdas en la columna
        for row_idx in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                # Convertir a string y calcular longitud
                cell_text = str(cell_value)
                # Considerar saltos de línea y caracteres especiales
                lines = cell_text.split('\n')
                max_line_length = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, max_line_length)
        
        # Ajustar ancho según el tipo de columna
        if 'ESPECIFICACIONES' in column_name.upper():
            # Para especificaciones, usar un ancho mínimo mayor
            calculated_width = max(max_length * 1.2, 50)
            worksheet.column_dimensions[col_letter].width = min(calculated_width, 80)  # Máximo 80
        elif 'MOTIVO' in column_name.upper():
            # Para motivos de bloqueo
            calculated_width = max(max_length * 1.2, 40)
            worksheet.column_dimensions[col_letter].width = min(calculated_width, 70)  # Máximo 70
        elif 'DESTINO' in column_name.upper():
            # Para destinos
            calculated_width = max(max_length * 1.2, 30)
            worksheet.column_dimensions[col_letter].width = min(calculated_width, 60)  # Máximo 60
        elif column_name.upper() in ['OBSERVACION', 'CARTONES', 'RESPONSABLE']:
            # Para campos más pequeños
            calculated_width = max(max_length * 1.2, 12)
            worksheet.column_dimensions[col_letter].width = min(calculated_width, 25)  # Máximo 25
        elif 'FECHA' in column_name.upper():
            # Para fechas, ancho fijo
            worksheet.column_dimensions[col_letter].width = 12
        elif 'CODIGO' in column_name.upper():
            # Para códigos
            worksheet.column_dimensions[col_letter].width = max(max_length * 1.2, 10)
        elif 'TALLA' in column_name.upper():
            # Para tallas
            worksheet.column_dimensions[col_letter].width = max(max_length * 1.2, 8)
        else:
            # Para otras columnas, usar cálculo automático con límites razonables
            calculated_width = max(max_length * 1.2, 15)
            worksheet.column_dimensions[col_letter].width = min(calculated_width, 50)  # Máximo 50
    
    # AJUSTE AUTOMÁTICO DE ALTURA DE FILAS
    for row_idx in range(1, worksheet.max_row + 1):
        max_lines = 1
        
        # Encontrar la celda con más líneas en esta fila
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                cell_text = str(cell_value)
                line_count = len(cell_text.split('\n'))
                # También considerar ajuste automático por longitud
                col_width = worksheet.column_dimensions[get_column_letter(col_idx)].width
                estimated_lines = max(1, len(cell_text) // int(col_width * 0.8))
                max_lines = max(max_lines, line_count, estimated_lines)
        
        # Ajustar altura de la fila
        if row_idx == 1:  # Encabezado
            worksheet.row_dimensions[row_idx].height = max(30, max_lines * 15)
        else:  # Datos
            worksheet.row_dimensions[row_idx].height = max(20, max_lines * 15)
    
    # Configuraciones finales
    worksheet.freeze_panes = 'A2'
    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

def guardar_excel_con_formato_mejorado(df_final, ruta_guardado):
    """Función mejorada para guardar Excel con formato automático"""
    try:
        # Crear el Excel con ExcelWriter para tener más control
        with pd.ExcelWriter(ruta_guardado, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Lotes con Novedades en Descongelado')
            
            # Aplicar formato mejorado
            format_excel_output_profesional(writer, df_final)
            
        print(f"\nReporte guardado exitosamente con formato automático:")
        print(f"{ruta_guardado}")
        return True
        
    except Exception as e: 
        print(f"Error al guardar con formato: {e}")
        
        # Intentar guardado básico sin formato
        try:
            df_final.to_excel(ruta_guardado.replace('.xlsx', '_sin_formato.xlsx'), index=False)
            print(f"Guardado sin formato en: {ruta_guardado.replace('.xlsx', '_sin_formato.xlsx')}")
            return False
        except Exception as e2:
            print(f"No se pudo guardar el archivo: {e2}")
            return False

def procesar_reporte(archivo_info, df_productos, df_tallas, df_defectos, fecha_inicio):
    """Procesa un reporte individual"""
    tipo_reporte, ruta_archivo = archivo_info
    
    try:
        df_reporte = pd.read_excel(ruta_archivo, engine='openpyxl', header=None, skiprows=6)
        df_reporte[4] = pd.to_datetime(df_reporte[4], errors='coerce')
        df_reporte.dropna(subset=[4, 5], inplace=True)
        df_filtrado = df_reporte[df_reporte[4] >= fecha_inicio].copy()
        
        print(f"Procesando {len(df_filtrado)} registros de {tipo_reporte}")
        
    except Exception as e:
        print(f"Error procesando {tipo_reporte}: {e}")
        return []

    reporte_data = []
    
    for index, fila in df_filtrado.iterrows():
        try:
            codigo_str = f"{int(fila.iloc[5]):06d}"
        except (ValueError, TypeError): 
            continue

        info_producto = df_productos[df_productos['CODIGO'] == codigo_str]
        if info_producto.empty: 
            continue
        
        all_motivos, all_specs = [], []
        
        m, s = validar_talla(fila, codigo_str, df_tallas, tipo_reporte); all_motivos.append(m); all_specs.append(s)
        m, s = validar_peso_bruto(fila, info_producto, tipo_reporte); all_motivos.append(m); all_specs.append(s)
        m, s = validar_peso_neto(fila, info_producto, tipo_reporte); all_motivos.append(m); all_specs.append(s)
        m, s = validar_uniformidad(fila, codigo_str, df_tallas, tipo_reporte); all_motivos.append(m); all_specs.append(s)
        m, s = validar_conteo_por_kg(fila, codigo_str, df_tallas, tipo_reporte); all_motivos.append(m); all_specs.append(s)
        motivos_def, specs_def = validar_defectos(fila, codigo_str, df_defectos, tipo_reporte)
        all_motivos.extend(motivos_def); all_specs.extend(specs_def)
        
        all_motivos = [i for i in all_motivos if i]
        all_specs = [i for i in all_specs if i]
        
        if all_motivos:
            destino_final = f"{info_producto['DESTINO_PAIS'].iloc[0]}/{info_producto['METODO_CONGELACION'].iloc[0]} {info_producto['CLIENTE'].iloc[0]} - {info_producto['EMBALAJE'].iloc[0]}"
            nueva_fila = {
                'fecha': fila.iloc[4].strftime('%d/%m/%Y'),       
                'codigo interno': fila.iloc[0],                    
                'talla': fila.iloc[8],                             
                'codigo': codigo_str,
                'destino': destino_final,
                'especificaciones de ficha tecnica': ' | '.join(all_specs),
                'motivo de bloqueo': ', '.join(all_motivos),
                'observacion': '', 
                'cartones': '', 
                'responsable': '',
                'tipo reporte': tipo_reporte
            }
            reporte_data.append(nueva_fila)
    
    return reporte_data

# --- Función Principal ---
def main():
    print("=== GENERADOR DE REPORTES: Lotes con Novedades en Descongelado ===\n")
    
    # Seleccionar archivos múltiples
    archivos_reporte = seleccionar_archivos_multiples()
    if not archivos_reporte: 
        print("Operacion cancelada.")
        return
    
    print(f"\nArchivos seleccionados: {len(archivos_reporte)}")
    for tipo, archivo in archivos_reporte:
        print(f"• {tipo}: {os.path.basename(archivo)}")
    
    # Seleccionar base de datos
    ruta_base_datos = seleccionar_archivo_excel("Selecciona tu Base de Datos (.xlsx)")
    if not ruta_base_datos: 
        print("Operacion cancelada.")
        return

    print("\nCargando base de datos...")
    try:
        df_productos = pd.read_excel(ruta_base_datos, sheet_name='PRODUCTOS', dtype={'CODIGO': str})
        df_tallas = pd.read_excel(ruta_base_datos, sheet_name='TALLAS', dtype={'CODIGO': str})
        df_defectos = pd.read_excel(ruta_base_datos, sheet_name='DEFECTOS', dtype={'CODIGO': str})
        for df in [df_productos, df_tallas, df_defectos]:
            df['CODIGO'] = df['CODIGO'].str.zfill(6)
        print("Base de datos cargada correctamente.")
    except Exception as e:
        print(f"Error al cargar base de datos: {e}")
        return

    # Obtener fechas máximas y configurar cada reporte
    configuraciones = {}
    for tipo_reporte, ruta_archivo in archivos_reporte:
        try:
            df_temp = pd.read_excel(ruta_archivo, engine='openpyxl', header=None, skiprows=6)
            df_temp[4] = pd.to_datetime(df_temp[4], errors='coerce')
            df_temp.dropna(subset=[4], inplace=True)
            if not df_temp.empty:
                fecha_max = df_temp[4].max()
                fecha_inicio = obtener_fecha_por_reporte(tipo_reporte, fecha_max)
                configuraciones[tipo_reporte] = {
                    'archivo': ruta_archivo,
                    'fecha_inicio': fecha_inicio,
                    'fecha_maxima': fecha_max
                }
        except Exception as e:
            print(f"Error procesando fechas de {tipo_reporte}: {e}")
            return

    # Procesar todos los reportes
    print(f"\n=== INICIANDO ANALISIS ===")
    todos_los_datos = []
    
    for tipo_reporte, config in configuraciones.items():
        archivo_info = (tipo_reporte, config['archivo'])
        datos_reporte = procesar_reporte(archivo_info, df_productos, df_tallas, df_defectos, config['fecha_inicio'])
        todos_los_datos.extend(datos_reporte)
        print(f"{tipo_reporte}: {len(datos_reporte)} lotes con novedades")

    # Generar reporte final
    if todos_los_datos:
        df_final = pd.DataFrame(todos_los_datos)
        df_final.columns = [col.upper() for col in df_final.columns]
        
        # Ordenar por fecha (más reciente primero) y tipo
        df_final['FECHA_SORT'] = pd.to_datetime(df_final['FECHA'], format='%d/%m/%Y')
        df_final = df_final.sort_values(['FECHA_SORT', 'TIPO REPORTE', 'CODIGO'], ascending=[False, True, True])
        df_final = df_final.drop('FECHA_SORT', axis=1)
        
        # Estadísticas finales
        total_lotes = len(df_final)
        entero_count = len(df_final[df_final['TIPO REPORTE'] == 'ENTERO'])
        cola_count = len(df_final[df_final['TIPO REPORTE'] == 'COLA'])
        
        print(f"\nRESUMEN FINAL")
        print(f"Total de lotes con novedades: {total_lotes}")
        if entero_count > 0:
            print(f"Lotes ENTERO: {entero_count}")
        if cola_count > 0:
            print(f"Lotes COLA: {cola_count}")
        
        # Guardar archivo con formato mejorado
        ruta_guardado = seleccionar_ruta_guardado()
        if ruta_guardado:
            success = guardar_excel_con_formato_mejorado(df_final, ruta_guardado)
            if not success:
                print("Se guardó una versión sin formato como respaldo.")
        else: 
            print("Guardado cancelado.")
    else:
        print("\nAnalisis completado. No se encontraron lotes con novedades.")

if __name__ == "__main__":
    main()