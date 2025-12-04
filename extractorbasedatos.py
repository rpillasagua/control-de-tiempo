import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import PyPDF2
import re
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class TechnicalSheetExtractor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Extractor de Fichas Técnicas - Versión DEFINITIVA")
        self.root.geometry("500x320")
        
        # Variables
        self.folder_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        
        self.setup_gui()
        
    def setup_gui(self):
        # Title
        title_label = tk.Label(self.root, text="Extractor de Fichas Técnicas", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=20)
        
        # Subtitle
        subtitle_label = tk.Label(self.root, text="Versión DEFINITIVA - DEFECTOS CORREGIDOS", 
                                 font=("Arial", 10, "italic"), fg="green")
        subtitle_label.pack(pady=(0,10))
        
        # Folder selection
        folder_frame = tk.Frame(self.root)
        folder_frame.pack(pady=10, padx=20, fill="x")
        
        tk.Label(folder_frame, text="Carpeta de PDFs:", font=("Arial", 10)).pack(anchor="w")
        
        folder_entry_frame = tk.Frame(folder_frame)
        folder_entry_frame.pack(fill="x", pady=5)
        
        tk.Entry(folder_entry_frame, textvariable=self.folder_path, width=50).pack(side="left", fill="x", expand=True)
        tk.Button(folder_entry_frame, text="Seleccionar", 
                 command=self.select_folder).pack(side="right", padx=(5,0))
        
        # Excel selection
        excel_frame = tk.Frame(self.root)
        excel_frame.pack(pady=10, padx=20, fill="x")
        
        tk.Label(excel_frame, text="Archivo Excel (Base de datos):", font=("Arial", 10)).pack(anchor="w")
        
        excel_entry_frame = tk.Frame(excel_frame)
        excel_entry_frame.pack(fill="x", pady=5)
        
        tk.Entry(excel_entry_frame, textvariable=self.excel_path, width=50).pack(side="left", fill="x", expand=True)
        tk.Button(excel_entry_frame, text="Seleccionar", 
                 command=self.select_excel).pack(side="right", padx=(5,0))

        
        # Process button
        tk.Button(self.root, text="Procesar Fichas Técnicas", 
                 command=self.process_files, font=("Arial", 12), 
                 bg="#4CAF50", fg="white", pady=10).pack(pady=30)
        
        # Status label
        self.status_label = tk.Label(self.root, text="Listo para procesar", 
                                   font=("Arial", 10), fg="blue")
        self.status_label.pack(pady=10)
        
        # Info label
        info_label = tk.Label(self.root, text="✓ PARSING CORREGIDO ✓ Sin concatenaciones ✓ 100% Funcional", 
                             font=("Arial", 8), fg="green")
        info_label.pack(pady=(0,10))
        
    def select_folder(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta con PDFs")
        if folder:
            self.folder_path.set(folder)
            
    def select_excel(self):
        excel_file = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*")]
        )
        if excel_file:
            self.excel_path.set(excel_file)
    
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF file"""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    except Exception as e:
                        print(f"Error extracting text from page: {e}")
                        continue
                return text
        except Exception as e:
            print(f"Error reading PDF {pdf_path}: {e}")
            return ""
            return ""

    def extract_code_from_text(self, text):
        """Extract code from PDF text, handling formats like FTPR-CPT-00038"""
        try:
            # Search for CÓDIGO: pattern
            match = re.search(r'CÓDIGO:\s*([\w-]+)', text, re.IGNORECASE)
            if match:
                full_code = match.group(1)
                # If it follows the pattern FTPR-CPT-XXXXX, extract the last digits
                if '-' in full_code:
                    parts = full_code.split('-')
                    # Check if the last part is numeric
                    if parts[-1].isdigit():
                        return parts[-1]
                return full_code
        except Exception:
            pass
        return None
        """Extract data between two patterns"""
        try:
            if end_pattern:
                pattern = f"{start_pattern}(.*?){end_pattern}"
                match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
            else:
                pattern = f"{start_pattern}(.*?)$"
                match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
            
            if match:
                return match.group(1).strip()
            return ""
        except Exception:
            return ""
            return ""

    def extract_section_data(self, text, start_marker, end_marker):
        """Extract text between two markers"""
        try:
            pattern = f"{start_marker}(.*?){end_marker}"
            match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
            if match:
                return match.group(1).strip()
            return ""
        except Exception:
            return ""

    def detect_product_type_from_text(self, text):
        """Infer product type from text content"""
        text_upper = text.upper()
        if 'ENTERO' in text_upper:
            return 'ENTERO'
        elif 'COLA' in text_upper:
            return 'COLA'
        elif 'VALOR AGREGADO' in text_upper:
            return 'VALOR AGREGADO'
        return 'DESCONOCIDO'

    def extract_technical_data(self, text, codigo, product_type):
        """Extract technical data from PDF text"""
        base_data = {
            'CODIGO': codigo,
            'VERSION': '',
            'DESCRIPCION': '',
            'CLIENTE': '',
            'MARCA': '',
            'METODO_CONGELACION': '',
            'CERTIFICACION': '',
            'COLOR': '',
            'EMBALAJE': '',
            'GLASEO_VALOR': '',
            'GLASEO_UNIDAD': '',
            'HIDRATANTE': '',
            'CONSERVANTE': '',
            'PESO_LIQUIDAR': '',
            'PESO_LIQUIDAR_UNIDAD': '',
            'PESO_NETO': '',
            'PESO_NETO_UNIDAD': '',
            'PESO_BRUTO_PRODUCCION': '',
            'PESO_BRUTO_PRODUCCION_UNIDAD': '',
            'PESO_BRUTO_MASTERS': '',
            'PESO_BRUTO_MASTERS_UNIDAD': '',
            'SOBREPESO_PORCENTAJE': '',
            'DESTINO_PAIS': '',
            'TIPO_PRODUCTO': product_type
        }
        
        try:
            clean_text = re.sub(r'\s+', ' ', text)
            
            # Extract basic information
            version_match = re.search(r'VERSIÓN:\s*(\d+)', clean_text, re.IGNORECASE)
            if version_match:
                base_data['VERSION'] = version_match.group(1)
            
            desc_match = re.search(r'DESCRIPCIÓN:\s*(.+?)(?=Cliente)', clean_text, re.IGNORECASE | re.DOTALL)
            if desc_match:
                base_data['DESCRIPCION'] = desc_match.group(1).strip()
            
            cliente_match = re.search(r'Cliente:\s*(.+?)(?=Destino)', clean_text, re.IGNORECASE | re.DOTALL)
            if cliente_match:
                base_data['CLIENTE'] = cliente_match.group(1).strip()
            
            # --- Código para extraer Destino/País ---
            pais_match = re.search(r'Destino/País\s*(.+?)(?=Marca)', clean_text, re.IGNORECASE | re.DOTALL)
            if pais_match:
                base_data['DESTINO_PAIS'] = pais_match.group(1).strip()
            # --- Fin del código nuevo ---
            
            marca_match = re.search(r'Marca\s+(.+?)(?=Especie)', clean_text, re.IGNORECASE | re.DOTALL)
            if marca_match:
                base_data['MARCA'] = marca_match.group(1).strip()
            
            metodo_match = re.search(r'Metodo de congelación\s+(.+?)(?=Vida)', clean_text, re.IGNORECASE | re.DOTALL)
            if metodo_match:
                base_data['METODO_CONGELACION'] = metodo_match.group(1).strip()
            
            cert_match = re.search(r'Certificación\s+(.+?)(?=Columna|Color)', clean_text, re.IGNORECASE | re.DOTALL)
            if cert_match:
                base_data['CERTIFICACION'] = cert_match.group(1).strip()
            
            color_match = re.search(r'Color\s+(.+?)(?=Pago|Origen)', clean_text, re.IGNORECASE | re.DOTALL)
            if color_match:
                base_data['COLOR'] = color_match.group(1).strip()
            
            embalaje_match = re.search(r'Embalaje\s+(.+?)(?=COMPLEMENTOS)', clean_text, re.IGNORECASE | re.DOTALL)
            if embalaje_match:
                raw_embalaje = embalaje_match.group(1).strip()
                # Intenta extraer el formato específico "10 Und * 2 Kg"
                clean_embalaje_match = re.search(r'(\d+\s*Und\s*\*\s*\d+(?:[.,]\d+)?\s*Kg)', raw_embalaje, re.IGNORECASE)
                if clean_embalaje_match:
                    base_data['EMBALAJE'] = clean_embalaje_match.group(1)
                else:
                    # Fallback: toma la primera línea o el texto completo si no encuentra el patrón
                    base_data['EMBALAJE'] = raw_embalaje.split('\n')[0].strip()
            
            # Extract glazing
            glaseo_patterns = [
                r'ML:\s*([0-9]+)\s*ml',
                r'GLASEO[:\s]*:\s*([0-9]+)%',
                r'GLASEO[:\s]*%?:\s*([0-9]+)',
                r'%:\s*([0-9]+)%'
            ]
            
            for i, pattern in enumerate(glaseo_patterns):
                glaseo_match = re.search(pattern, clean_text, re.IGNORECASE)
                if glaseo_match:
                    base_data['GLASEO_VALOR'] = glaseo_match.group(1)
                    if i == 0:
                        base_data['GLASEO_UNIDAD'] = 'ml'
                    else:
                        base_data['GLASEO_UNIDAD'] = '%'
                    break
            
            # Extract conservante
            conservante_match = re.search(r'Conservante\s+(.+?)(?=HIDRATANTE|$)', clean_text, re.IGNORECASE | re.DOTALL)
            if conservante_match:
                base_data['CONSERVANTE'] = conservante_match.group(1).strip()
            
            # Extract peso information
            peso_liquidar_match = re.search(r'Peso a liquidar \((\w+)\):\s*([0-9.]+)', clean_text, re.IGNORECASE)
            if peso_liquidar_match:
                base_data['PESO_LIQUIDAR_UNIDAD'] = peso_liquidar_match.group(1).upper()
                base_data['PESO_LIQUIDAR'] = peso_liquidar_match.group(2)
            
            peso_neto_match = re.search(r'Peso Neto declarado \((\w+)\):\s*([0-9.]+)', clean_text, re.IGNORECASE)
            if peso_neto_match:
                base_data['PESO_NETO_UNIDAD'] = peso_neto_match.group(1).upper()
                base_data['PESO_NETO'] = peso_neto_match.group(2)
            
            peso_bruto_prod_match = re.search(r'Peso Bruto Producción \((\w+)\):\s*([0-9.]+)', clean_text, re.IGNORECASE)
            if peso_bruto_prod_match:
                base_data['PESO_BRUTO_PRODUCCION_UNIDAD'] = peso_bruto_prod_match.group(1).upper()
                base_data['PESO_BRUTO_PRODUCCION'] = peso_bruto_prod_match.group(2)
            
            peso_bruto_masters_match = re.search(r'Peso Bruto masters \((\w+)\)\s*([0-9.]+)', clean_text, re.IGNORECASE)
            if peso_bruto_masters_match:
                base_data['PESO_BRUTO_MASTERS_UNIDAD'] = peso_bruto_masters_match.group(1).upper()
                base_data['PESO_BRUTO_MASTERS'] = peso_bruto_masters_match.group(2)
            
            sobrepeso_match = re.search(r'Sobrepeso \(%\):\s*([0-9.]+%?)', clean_text, re.IGNORECASE)
            if sobrepeso_match:
                base_data['SOBREPESO_PORCENTAJE'] = sobrepeso_match.group(1)
            
            hidratante_match = re.search(r'Lleva hidratante:\s*(SI|NO)', clean_text, re.IGNORECASE)
            if hidratante_match:
                base_data['HIDRATANTE'] = hidratante_match.group(1).upper()
            
        except Exception as e:
            print(f"Error extracting data from {codigo}: {e}")
        
        return base_data
    
    def extract_sizes_data_normalized(self, text, codigo):
        """Extract sizes data in normalized format"""
        sizes_data = []
        
        tallas_section = self.extract_section_data(text, "TALLAS", "DEFECTOS")
        if not tallas_section:
            return sizes_data
        
        lines = tallas_section.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or any(header in line.upper() for header in ['TALLA DE MP', 'CONTEO DE MP', 'TALLA MARCADA', 'CONTEO FINAL', 'UNIFORMIDAD', 'CODIGO DE', 'BARRAS', 'PRODUCTO', 'OBSERVACIÓN']):
                continue
            
            parts = [part for part in line.split() if part.strip()]
            
            # Necesitamos al menos 6 partes para tener datos completos, 
            # pero con nombres largos pueden ser más partes.
            # Estrategia: Parsear de DERECHA a IZQUIERDA las últimas 4 columnas que suelen ser fijas.
            
            if len(parts) >= 5: # Mínimo para tener TallaMP+ConteoMP (1) + 4 columnas fijas
                try:
                    # 1. Extraer las últimas 4 columnas (asumiendo que son tokens únicos)
                    uniformidad = parts[-1]
                    talla_marcada = parts[-2]
                    conteo_final = parts[-3]
                    talla_venta = parts[-4]
                    
                    # 2. El resto es la parte izquierda (Talla MP + Conteo MP)
                    left_parts = parts[:-4]
                    left_text = " ".join(left_parts)
                    
                    # 3. Separar Talla MP y Conteo MP
                    # Buscamos el punto donde empieza el Conteo MP.
                    # Heurística: El Conteo MP suele empezar con un dígito o 'U' (ej: 71/90, 200, U7)
                    # O casos especiales como "200 EN ADELANTE"
                    
                    talla_mp = ""
                    conteo_mp = ""
                    
                    # Regex para encontrar el inicio del conteo (Digito o U seguido de digito)
                    # Buscamos la primera palabra que coincida con patrón de conteo
                    split_index = -1
                    first_match_index = -1
                    
                    for i, part in enumerate(left_parts):
                        if re.match(r'^(U?\d)', part, re.IGNORECASE):
                            first_match_index = i
                            break
                    
                    if first_match_index != -1:
                        # Si la coincidencia es en el índice 0, intentamos encontrar otra más adelante
                        # Esto maneja casos donde Talla MP parece un número (ej: U7-31/35)
                        if first_match_index == 0 and len(left_parts) > 1:
                            second_match_index = -1
                            for i in range(1, len(left_parts)):
                                if re.match(r'^(U?\d)', left_parts[i], re.IGNORECASE):
                                    second_match_index = i
                                    break
                            
                            if second_match_index != -1:
                                split_index = second_match_index
                            else:
                                split_index = 0
                        else:
                            split_index = first_match_index
                    
                    if split_index != -1:
                        talla_mp = " ".join(left_parts[:split_index])
                        conteo_mp = " ".join(left_parts[split_index:])
                    else:
                        # Si no encontramos patrón claro, asumimos que todo es Talla MP (o error)
                        # O dividimos a la mitad si hay 2 partes? No, mejor todo a Talla MP y warning
                        talla_mp = left_text
                        conteo_mp = "N/A"

                    # Limpieza final
                    if not talla_mp: talla_mp = "N/A"
                    
                    if not all(part == 'N/A' for part in [talla_mp, conteo_mp, talla_venta, conteo_final, uniformidad]):
                        sizes_data.append({
                            'CODIGO': codigo,
                            'TALLA_MP': talla_mp,
                            'CONTEO_MP': conteo_mp,
                            'TALLA_VENTA': talla_venta,
                            'TALLA_MARCADA': talla_marcada,
                            'CONTEO_FINAL': conteo_final,
                            'UNIFORMIDAD': uniformidad
                        })
                except Exception as e:
                    print(f"Error parseando linea de tallas '{line}': {e}")
        
        return sizes_data
    
    def extract_defects_from_table(self, text, codigo, product_type):
        """
        Extrae TODOS los defectos de la tabla de DEFECTOS - VERSIÓN FINAL (Busca Múltiples Defectos por Línea)
        """
        defects_data = []
        
        defectos_section = self.extract_section_data(text, "DEFECTOS", "REQUISITOS")
        if not defectos_section:
            print(f"No se encontró la sección DEFECTOS para el código {codigo}")
            return defects_data

        print(f"\n=== PROCESANDO DEFECTOS PARA {codigo} ===")
        
        # LISTA MAESTRA DEFINITIVA DE DEFECTOS
        known_defects = [
            'RESIDUOS DE HEPATOPANCREAS', 'BRANQUIAS AMARILLAS FUERTES', 'BRANQUIAS AMARILLAS LEVES',
            'BRANQUIAS OSCURAS FUERTES', 'BRANQUIAS OSCURAS LEVES', 'SABOR GALLINAZA FUERTE',
            'SABOR GALLINAZA LEVE', 'MANCHAS NEGRAS FUERTES', 'HONGO BUCAL FUERTE',
            'SABOR TIERRA FUERTE', 'MANCHAS NEGRAS LEVES', 'HONGO BUCAL LEVE',
            'SABOR TIERRA LEVE', 'RESIDUAL DE SULFITO', 'VARIACION DE COLOR',
            'CABEZA DESCOLGADA', 'MAL DESCABEZADO', 'DEFORMES FUERTES',
            'HEMOLINFAS LEVES', 'HEPATO REVENTADO', 'MATERIAL EXTRAÑO', 'CASCARA APARTE',
            'CORTE IRREGULAR', 'HEMOLINFAS FUERTES', 'DEFORMES LEVES', 'DESHIDRATADO',
            'HEPATO REGADO', 'RESTOS DE VENAS', 'CABEZA FUERTE', 'CABEZA NARANJA',
            'CORTE PROFUNDO', 'FALTA DE CORTE', 'DEFECTOS TOTALES', 'CABEZA FLOJA',
            'CORTE LARGO', 'MALTRATADO', 'SEMIROSADO', 'SIN TELSON', 'CABEZA ROJA',
            'SABOR CHOCLO', 'SABOR COMBUSTIBLE', 'QUEBRADO', 'MELANOSIS', 'FLACIDO',
            'MUDADO', 'ROSADO', 'COLOR', 'SABOR'
        ]
        
        # Patrón para limpiar y extraer solo el valor relevante de una cadena
        value_cleaning_pattern = re.compile(
            r'('
            r'NO\s*ACEPTA\s*VARIACIÓN|'
            r'\d{1,3}(?:[.,]\d+)?\s*%|'
            r'\d+\s*-\s*\d+\s*ppm|'
            r'SI|NO|N/A|'
            r'A[1-4](?:,\s*A[1-4])*|'
            r'AUSENCIA|PRESENCIA'
            r')',
            re.IGNORECASE
        )
        
        # Unir todos los defectos conocidos en un gran patrón de búsqueda
        # Se escapan caracteres especiales para que funcionen en el regex
        defects_pattern = re.compile(
            '|'.join(re.escape(defect) for defect in known_defects),
            re.IGNORECASE
        )
        
        lines = defectos_section.strip().split('\n')
        
        for line in lines:
            # Limpieza exhaustiva de la línea
            cleaned_line = re.sub(r'[:]', '', line).strip()
            if not cleaned_line:
                continue
            
            # 1. Encontrar todas las ocurrencias de defectos conocidos en la línea
            found_defects = list(defects_pattern.finditer(cleaned_line))
            
            # 2. Iterar sobre los defectos encontrados para asignarles su valor
            for i, match in enumerate(found_defects):
                defect_name = match.group(0)
                
                # El valor es el texto entre el final de este defecto y el inicio del siguiente
                start_of_value = match.end()
                end_of_value = found_defects[i+1].start() if i + 1 < len(found_defects) else len(cleaned_line)
                
                value_part = cleaned_line[start_of_value:end_of_value].strip()
                
                # Limpiar el valor extraído para obtener solo la parte importante
                value_match = value_cleaning_pattern.search(value_part)
                
                final_value = value_part # Por defecto, usar el texto restante
                if value_match:
                    final_value = value_match.group(0).strip() # Si hay coincidencia, usar el valor limpio
                
                if final_value: # Solo añadir si se encontró un valor
                    defects_data.append({
                        'CODIGO': codigo,
                        'TIPO_PRODUCTO': product_type,
                        'DEFECTO': defect_name.strip().upper().replace(' ', '_'),
                        'VALOR': final_value
                    })

        print(f"=== SE EXTRAJERON {len(defects_data)} DEFECTOS para el código {codigo} ===")
        return defects_data

    def create_normalized_excel_report(self, all_products_data, all_sizes_data, all_defects_data, output_path):
        """Create Excel report with 3 sheets"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # PRODUCTOS sheet
            if all_products_data:
                products_df = pd.DataFrame(all_products_data)
                products_df = products_df.loc[:, (products_df != '').any(axis=0)]
                products_df.to_excel(writer, sheet_name='PRODUCTOS', index=False)
                self.format_sheet(writer, 'PRODUCTOS')
            
            # TALLAS sheet
            if all_sizes_data:
                sizes_df = pd.DataFrame(all_sizes_data)
                sizes_df.to_excel(writer, sheet_name='TALLAS', index=False)
                self.format_sheet(writer, 'TALLAS')
            else:
                empty_sizes_df = pd.DataFrame(columns=['CODIGO', 'TALLA_MP', 'CONTEO_MP', 'TALLA_VENTA', 'TALLA_MARCADA', 'CONTEO_FINAL', 'UNIFORMIDAD'])
                empty_sizes_df.to_excel(writer, sheet_name='TALLAS', index=False)
                self.format_sheet(writer, 'TALLAS')
            
            # DEFECTOS sheet
            if all_defects_data:
                defects_df = pd.DataFrame(all_defects_data)
                defects_df.to_excel(writer, sheet_name='DEFECTOS', index=False)
                self.format_sheet(writer, 'DEFECTOS')
            else:
                empty_defects_df = pd.DataFrame(columns=['CODIGO', 'TIPO_PRODUCTO', 'DEFECTO', 'VALOR'])
                empty_defects_df.to_excel(writer, sheet_name='DEFECTOS', index=False)
                self.format_sheet(writer, 'DEFECTOS')
    
    def format_sheet(self, writer, sheet_name):
        """Apply formatting to Excel sheet"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        worksheet.row_dimensions[1].height = 25
        
        # Auto-adjust column widths
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 3, 10), 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"Warning: Could not auto-adjust column widths: {e}")
        
        # Alternating row colors
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_num in range(2, worksheet.max_row + 1):
            if row_num % 2 == 0:
                for cell in worksheet[row_num]:
                    if cell.fill.start_color.index == '00000000':
                        cell.fill = light_fill
    
    def load_excel_database(self, excel_path):
        """Load Excel database"""
        try:
            print(f"Cargando base de datos desde: {excel_path}")
            
            # Estrategia 1: Intentar leer por nombres de columna (Header en fila 7/Index 6)
            try:
                df = pd.read_excel(excel_path, header=6)
                # Normalizar columnas
                df.columns = [str(c).upper().strip() for c in df.columns]
            except:
                df = pd.DataFrame()

            has_names = 'CODIGO' in df.columns and 'TIPO_PRODUCTO' in df.columns
            
            # Estrategia 2: Si no hay nombres, usar posiciones fijas (Usuario: Col B=Codigo, Col G=Tipo, Datos fila 8)
            if not has_names:
                print("Formato por nombre no detectado. Usando formato posicional (Col B y G)...")
                # skiprows=7 para empezar datos en fila 8
                df = pd.read_excel(excel_path, header=None, skiprows=7)
                
                if len(df.columns) > 6:
                    # Renombrar columnas clave: 1 (B) -> CODIGO, 6 (G) -> TIPO_PRODUCTO
                    df = df.rename(columns={1: 'CODIGO', 6: 'TIPO_PRODUCTO'})
                else:
                    raise Exception("El archivo Excel no tiene suficientes columnas (se requiere columna G)")
            
            # Procesar datos
            if 'CODIGO' not in df.columns or 'TIPO_PRODUCTO' not in df.columns:
                 raise Exception("No se pudieron identificar las columnas CODIGO y TIPO_PRODUCTO")

            df = df.dropna(subset=['CODIGO'])
            
            code_to_type = {}
            for _, row in df.iterrows():
                try:
                    # Procesar Código
                    raw_code = str(row['CODIGO']).strip()
                    if '.' in raw_code: # Eliminar decimales si es numero (ej 38.0)
                        raw_code = raw_code.split('.')[0]
                    
                    # Limpiar código (solo números)
                    code_clean = ''.join(filter(str.isdigit, raw_code))
                    
                    # Procesar Tipo
                    product_type = str(row['TIPO_PRODUCTO']).strip().upper()
                    
                    if code_clean:
                        # Guardar variaciones para asegurar match
                        code_padded = code_clean.zfill(5)
                        
                        code_to_type[code_clean] = product_type
                        code_to_type[code_padded] = product_type
                        code_to_type[raw_code] = product_type
                        
                except Exception:
                    continue
            
            print(f"Base de datos cargada exitosamente. {len(code_to_type)} códigos indexados.")
            return code_to_type
            
        except Exception as e:
            print(f"Error detallado cargando Excel: {str(e)}")
            raise Exception(f"Error cargando base de datos: {str(e)}")
    
    def process_files(self):
        if not self.folder_path.get() or not self.excel_path.get():
            messagebox.showerror("Error", "Por favor seleccione la carpeta de PDFs y el archivo Excel")
            return
        
        try:
            self.status_label.config(text="Cargando base de datos Excel...", fg="orange")
            self.root.update()
            
            code_to_type = self.load_excel_database(self.excel_path.get())

            
            self.status_label.config(text="Procesando PDFs...", fg="orange")
            self.root.update()
            
            pdf_folder = Path(self.folder_path.get())
            
            all_products_data = []
            all_sizes_data = []
            all_defects_data = []
            
            # Get all PDF files
            pdf_files = [f for f in os.listdir(self.folder_path.get()) if f.lower().endswith('.pdf')]
            
            # Sort files to ensure organized output
            pdf_files.sort()
            
            total_files = len(pdf_files)
            print(f"Se encontraron {total_files} archivos PDF en la carpeta seleccionada.")
            
            if total_files == 0:
                messagebox.showwarning("Advertencia", "No se encontraron archivos PDF en la carpeta seleccionada.")
                return
            
            processed_files = 0
            for pdf_file in pdf_files:
                codigo = os.path.splitext(pdf_file)[0]
                codigo_padded = str(codigo).zfill(5)
                
                self.status_label.config(
                    text=f"Procesando {codigo} ({processed_files + 1}/{total_files})...", 
                    fg="orange"
                )
                self.root.update()
                
                full_pdf_path = os.path.join(self.folder_path.get(), pdf_file)
                text = self.extract_text_from_pdf(full_pdf_path)
                
                # Try to extract code from text first
                extracted_code = self.extract_code_from_text(text)
                if extracted_code:
                    codigo = extracted_code
                    codigo_padded = str(codigo).zfill(5)
                
                # Look up product type with the (possibly new) code
                product_type = code_to_type.get(codigo_padded, 'UNKNOWN')
                if product_type == 'UNKNOWN':
                    product_type = code_to_type.get(codigo, 'UNKNOWN')
                
                # Fallback: Try to detect type from text if not found in Excel
                if product_type == 'UNKNOWN' and text:
                    product_type = self.detect_product_type_from_text(text)
                    print(f"Tipo no encontrado en Excel para {codigo}. Detectado por texto: {product_type}")


                
                if text:
                    try:
                        print(f"   Extrayendo datos tecnicos...")
                        product_data = self.extract_technical_data(text, codigo, product_type)
                        all_products_data.append(product_data)
                        
                        print(f"   Extrayendo tallas...")
                        sizes_data = self.extract_sizes_data_normalized(text, codigo)
                        all_sizes_data.extend(sizes_data)
                        
                        print(f"   Extrayendo defectos...")
                        defects_data = self.extract_defects_from_table(text, codigo, product_type)
                        all_defects_data.extend(defects_data)
                        
                    except Exception as e:
                        print(f"ERROR procesando archivo {codigo}: {str(e)}")
                        # Continue to next file instead of stopping
                        continue
                
                processed_files += 1
                if processed_files % 10 == 0:
                    print(f"Procesados {processed_files}/{total_files} archivos...")
            
            self.status_label.config(text="Generando reporte Excel...", fg="orange")
            self.root.update()
            
            output_path = pdf_folder / "Base_Datos_Fichas_Tecnicas_DEFINITIVA.xlsx"
            self.create_normalized_excel_report(all_products_data, all_sizes_data, all_defects_data, output_path)
            
            summary = f"""¡Procesamiento completado!

Archivos procesados: {processed_files}
Productos extraídos: {len(all_products_data)}
Registros de tallas: {len(all_sizes_data)}
Registros de defectos: {len(all_defects_data)}

Archivo guardado: {output_path}

MEJORAS IMPLEMENTADAS:
✓ Parsing de defectos completamente corregido
✓ Sin concatenaciones erróneas
✓ Manejo correcto de defectos LEVE/FUERTE
✓ Detección mejorada de valores compuestos (ppm, %, etc.)
✓ Estructura normalizada mantenida"""
            
            self.status_label.config(text="¡Completado!", fg="green")
            messagebox.showinfo("Éxito", summary)
            
        except Exception as e:
            self.status_label.config(text="Error en procesamiento", fg="red")
            messagebox.showerror("Error", f"Error: {str(e)}")
    
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    try:
        import PyPDF2
        import openpyxl
    except ImportError:
        print("Installing required packages...")
        import subprocess
        import sys
        
        packages = ['PyPDF2', 'openpyxl', 'pandas']
        for package in packages:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        
        import PyPDF2
        import openpyxl
    
    app = TechnicalSheetExtractor()
    app.run()